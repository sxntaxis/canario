"""Deterministic, format-aware workbook materialization for table evidence."""

from __future__ import annotations

import datetime as _datetime
import hashlib
import json
from dataclasses import dataclass

from openpyxl import load_workbook

from .contracts import DerivativeOutput, ProcessorDescriptor, ProcessorInvocation, ProcessorResult, QualitySignal


def _json_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, bool):
        return {"type": "boolean", "value": value}
    if isinstance(value, int) and not isinstance(value, bool):
        return {"type": "integer", "value": value}
    if isinstance(value, float):
        return {"type": "number", "value": value}
    if isinstance(value, (_datetime.datetime, _datetime.date, _datetime.time)):
        return {"type": "datetime", "value": value.isoformat()}
    if isinstance(value, str):
        return {"type": "string", "value": value}
    return {"type": "string", "value": str(value)}


@dataclass(frozen=True, slots=True)
class StructuredTableConfig:
    max_input_bytes: int = 256 * 1024 * 1024
    max_scopes: int = 512


class StructuredTableProcessor:
    """Preserve workbook structure without interpreting document vocabulary."""

    def __init__(self, *, config: StructuredTableConfig | None = None) -> None:
        self.config = config or StructuredTableConfig()
        self._descriptor = ProcessorDescriptor(
            key="core.xlsx_table",
            capability_key="structured_table",
            implementation_version="openpyxl-3.1.5",
            execution_venue="local_deterministic",
            input_media_types=frozenset({
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }),
            output_kinds=frozenset({"table"}),
            scope_kinds=frozenset({"whole"}),
            max_input_bytes=self.config.max_input_bytes,
            max_scopes=self.config.max_scopes,
        )

    @property
    def descriptor(self) -> ProcessorDescriptor:
        return self._descriptor

    def process(self, invocation: ProcessorInvocation) -> ProcessorResult:
        if invocation.media_type not in self.descriptor.input_media_types:
            return ProcessorResult("failed", error_code="unsupported_media_type")
        if len(invocation.source_bytes) > self.config.max_input_bytes:
            return ProcessorResult("failed", error_code="input_too_large")
        try:
            workbook = load_workbook(invocation_source(invocation), data_only=False, read_only=False)
            sheets = []
            for ordinal, worksheet in enumerate(workbook.worksheets, start=1):
                rows = []
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=worksheet.max_row,
                    min_col=1,
                    max_col=worksheet.max_column,
                ):
                    rows.append([
                        {
                            "address": cell.coordinate,
                            "value": (
                                {"type": "formula", "value": cell.value}
                                if cell.data_type == "f" else _json_value(cell.value)
                            ),
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                        }
                        for cell in row
                    ])
                sheets.append({
                    "name": worksheet.title,
                    "ordinal": ordinal,
                    "state": worksheet.sheet_state,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "merged_ranges": sorted(str(value) for value in worksheet.merged_cells.ranges),
                    "rows": rows,
                })
            payload = {
                "format": "canario.structured_table.v1",
                "source_sha256": hashlib.sha256(invocation.source_bytes).hexdigest(),
                "sheets": sheets,
            }
            data = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
            evidence = tuple(
                QualitySignal(scope.id, "table.sheet_count", "v1", len(sheets))
                for scope in invocation.scopes
            )
            return ProcessorResult(
                "success",
                (DerivativeOutput(data, "table", "application/json", charset="utf-8"),),
                evidence,
            )
        except Exception:
            return ProcessorResult("failed", error_code="workbook_parse_failed")


def invocation_source(invocation: ProcessorInvocation):
    """Provide a seekable stream while keeping the processor byte-authoritative."""
    import io

    return io.BytesIO(invocation.source_bytes)
