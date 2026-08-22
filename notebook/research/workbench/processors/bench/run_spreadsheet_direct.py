"""Record a bounded direct-structured XLSX inspection for the civic bench."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--result", type=Path, required=True)
    args = parser.parse_args()
    workbook = load_workbook(args.input, data_only=False, read_only=True)
    displayed = load_workbook(args.input, data_only=True, read_only=True)
    sheets = []
    for sheet, displayed_sheet in zip(workbook.worksheets, displayed.worksheets):
        formulas = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.coordinate)
        selected = {
            "A1": sheet["A1"].value,
            "C1": sheet["C1"].value,
            "O1": sheet["O1"].value,
            "A2": sheet["A2"].value,
            "C2": sheet["C2"].value,
            "O2_formula_or_value": sheet["O2"].value,
            "O2_displayed_value": displayed_sheet["O2"].value,
        }
        sheets.append(
            {
                "title": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "formula_count": len(formulas),
                "formula_cells_sample": formulas[:10],
                "selected_cells": selected,
            }
        )
    result = {
        "bench_state": "CIVIC_PROCESSOR_BENCH_PARTIAL__XLSX_DIRECT_STRUCTURED",
        "source_sha256": sha256(args.input),
        "source_bytes": args.input.stat().st_size,
        "parser": "openpyxl",
        "parser_version": "3.1.5",
        "license": "CC BY",
        "sheets": sheets,
        "architectural_conclusion": "XLSX enters through a direct structured representation path; OCR/Codex escalation is not the default for workbook cells.",
    }
    args.result.parent.mkdir(parents=True, exist_ok=True)
    args.result.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
